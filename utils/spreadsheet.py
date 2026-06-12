"""
utils/spreadsheet.py — builds a formatted Excel workbook (.xlsx) from scan results.

Sheets:
  1. Top Picks        — top 5 bull + top 5 bear with full conviction details
  2. All Results      — every scanned ticker sorted by net score
  3. Signal Detail    — per-signal breakdown for every ticker
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Tuple

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

import config
from signals.base import TickerAnalysis
from signals.conviction import ConvictionScore

log = logging.getLogger(__name__)

# ── Color palette ─────────────────────────────────────────────────────────────
BULL_DARK   = "1D6F42"
BULL_MID    = "2D9E5F"
BULL_LIGHT  = "C6EFCE"
BEAR_DARK   = "9C0006"
BEAR_MID    = "C82333"
BEAR_LIGHT  = "FFC7CE"
NEUTRAL_BG  = "F2F2F2"
HEADER_BG   = "1F3864"
HEADER_FG   = "FFFFFF"
ACCENT_BG   = "2E75B6"
ACCENT_FG   = "FFFFFF"
BORDER_CLR  = "BFBFBF"

from signals import SIG_NAMES as _SIG_NAMES_FULL
# Short names for column headers (truncated for space)
SIG_NAMES = [s.split(".")[0].split(" ")[0][:6] for s in _SIG_NAMES_FULL]


def _thin_border() -> Border:
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_font(bold: bool = True, color: str = HEADER_FG, size: int = 10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _body_font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _write_header_row(ws, row: int, cols: List[str], fill_color: str = HEADER_BG) -> None:
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font      = _header_font()
        cell.fill      = _fill(fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()


def _bias_fill(bias_str: str) -> PatternFill:
    if bias_str == "bull":    return _fill(BULL_LIGHT)
    if bias_str == "bear":    return _fill(BEAR_LIGHT)
    return _fill(NEUTRAL_BG)


def _score_fill(score: int) -> PatternFill:
    if score >= 4:  return _fill(BULL_LIGHT)
    if score >= 2:  return _fill("E2EFDA")
    if score <= -4: return _fill(BEAR_LIGHT)
    if score <= -2: return _fill("FCE4D6")
    return _fill(NEUTRAL_BG)


def _set_col_widths(ws, widths: dict) -> None:
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ── Sheet 1: Top Picks ────────────────────────────────────────────────────────

def _sheet_top_picks(wb, bulls: List[Tuple], bears: List[Tuple], universe: str, ts: str) -> None:
    ws = wb.active
    ws.title = "Top Picks"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value     = f"Trading Signal Scanner — Top Picks  |  Universe: {universe.upper()}  |  {ts}"
    t.font      = Font(name="Calibri", bold=True, size=13, color=HEADER_FG)
    t.fill      = _fill(HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    row = 3
    for section_label, section_fill, picks in [
        ("▲  TOP 5 BULLISH — Highest Conviction Buys", BULL_DARK, bulls),
        ("▼  TOP 5 BEARISH — Highest Conviction Shorts / Avoids", BEAR_DARK, bears),
    ]:
        ws.merge_cells(f"A{row}:L{row}")
        c = ws.cell(row=row, column=1, value=section_label)
        c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill      = _fill(section_fill)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        cols = ["Rank", "Ticker", "Company", "Price", "Chg %", "Score", "Bull", "Bear",
                "Grade", "Conviction %", "Verdict", "Key Signals", "Analysis",
                "Fib Direction", "Anchor", "Entry (38.2%)", "Stop (61.8%)",
                "Target 1 (100%)", "Target 2 (127.2%)", "Target 3 (161.8%)",
                "R/R T1", "R/R T2", "R/R T3", "Next-Hr Target"]
        _write_header_row(ws, row, cols, ACCENT_BG)
        ws.row_dimensions[row].height = 30
        row += 1

        for rank, (ta, cs) in enumerate(picks, 1):
            key_str  = " | ".join(cs.key_signals[:3])
            fib = ta.fib
            def _fp(v): return f"${v:.2f}" if v else "—"
            row_data = [
                rank,
                ta.ticker,
                ta.company_name,
                f"${ta.price:.2f}" if ta.price else "—",
                f"{ta.chg_pct:+.2f}%" if ta.chg_pct else "—",
                f"{ta.net_score:+d}",
                ta.bull_count,
                ta.bear_count,
                cs.grade,
                f"{cs.conviction_pct:.1f}%",
                ta.verdict,
                key_str,
                cs.analysis,
                fib.direction.upper()     if fib else "—",
                fib.anchor_type           if fib else "—",
                _fp(fib.entry_price)      if fib else "—",
                _fp(fib.stop_loss)        if fib else "—",
                _fp(fib.target_1)         if fib else "—",
                _fp(fib.target_2)         if fib else "—",
                _fp(fib.target_3)         if fib else "—",
                f"{fib.risk_reward_t1:.1f}x" if (fib and fib.risk_reward_t1) else "—",
                f"{fib.risk_reward_t2:.1f}x" if (fib and fib.risk_reward_t2) else "—",
                f"{fib.risk_reward_t3:.1f}x" if (fib and fib.risk_reward_t3) else "—",
                _fp(fib.next_hour_target) if fib else "—",
            ]
            sfill = _fill(BULL_LIGHT) if cs.direction == "bullish" else _fill(BEAR_LIGHT)
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=c_idx, value=val)
                cell.font      = _body_font(bold=(c_idx == 2))
                cell.fill      = sfill
                cell.border    = _thin_border()
                cell.alignment = Alignment(
                    vertical="top", wrap_text=(c_idx == 12),
                    horizontal="center" if c_idx in (1,4,5,6,7,8,9) else "left"
                )
            ws.row_dimensions[row].height = 70 if len(cs.analysis) > 200 else 50
            row += 1
        row += 2

    _set_col_widths(ws, {
        "A": 6,  "B": 9,  "C": 26, "D": 10, "E": 9,  "F": 8,
        "G": 6,  "H": 6,  "I": 8,  "J": 12, "K": 18,
        "L": 36, "M": 72, "N": 10, "O": 22, "P": 14,
        "Q": 14, "R": 14, "S": 15, "T": 15, "U": 9,
        "V": 9,  "W": 9,  "X": 14,
    })
    ws.freeze_panes = "A3"


# ── Sheet 2: All Results ──────────────────────────────────────────────────────

def _sheet_all_results(wb, results: List[TickerAnalysis], universe: str, ts: str) -> None:
    ws = wb.create_sheet("All Results")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value     = f"All Scanned Tickers — {universe.upper()} — {ts}  ({len(results)} tickers)"
    t.font      = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
    t.fill      = _fill(HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    cols = ["Ticker", "Company", "Price", "Chg %", "Net Score", "Bull", "Bear",
            "Verdict", "Bars", "Mode"] + SIG_NAMES
    _write_header_row(ws, 2, cols)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    for i, ta in enumerate(results, 3):
        row_vals = [
            ta.ticker,
            ta.company_name,
            round(ta.price, 2) if ta.price else None,
            round(ta.chg_pct, 2) if ta.chg_pct else None,
            ta.net_score,
            ta.bull_count,
            ta.bear_count,
            ta.verdict,
            ta.bars,
            ta.mode,
        ] + [s.bias.value for s in ta.signals]

        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=c_idx, value=val)
            cell.font      = _body_font(bold=(c_idx == 1))
            cell.border    = _thin_border()
            cell.alignment = Alignment(
                horizontal="left" if c_idx in (1, 2) else "center",
                vertical="center"
            )
            # score coloring — shift index by 1 for new Company col
            if c_idx == 5:
                cell.fill = _score_fill(ta.net_score)
            elif c_idx >= 11:
                cell.fill = _bias_fill(str(val))
            elif i % 2 == 0:
                cell.fill = _fill("F9F9F9")

        # % format
        ws.cell(row=i, column=4).number_format = '+0.00%;-0.00%'

    _set_col_widths(ws, {
        "A": 9,  "B": 26, "C": 9, "D": 10, "E": 10, "F": 7,
        "G": 7,  "H": 20, "I": 8,  "J": 9,
        "K": 9,  "L": 9,  "M": 9,  "N": 9,
        "O": 9,  "P": 9,  "Q": 9,
    })


# ── Sheet 3: Signal Detail ────────────────────────────────────────────────────

def _sheet_signal_detail(wb, results: List[TickerAnalysis]) -> None:
    ws = wb.create_sheet("Signal Detail")
    ws.sheet_view.showGridLines = False

    header = ["Ticker", "Price", "Score"]
    for n in SIG_NAMES:
        header += [n, f"{n} label"]
    _write_header_row(ws, 1, header)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for i, ta in enumerate(results, 2):
        row_vals = [ta.ticker, round(ta.price, 2) if ta.price else None, ta.net_score]
        for s in ta.signals:
            row_vals += [s.bias.value, s.label]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=c_idx, value=val)
            cell.font      = _body_font()
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if c_idx == 3:
                cell.fill = _score_fill(ta.net_score)
            elif c_idx >= 4 and (c_idx - 4) % 2 == 0:
                col_sig_idx = (c_idx - 4) // 2
                if col_sig_idx < len(ta.signals):
                    cell.fill = _bias_fill(ta.signals[col_sig_idx].bias.value)

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    for i in range(4, 4 + len(SIG_NAMES) * 2, 2):
        ws.column_dimensions[get_column_letter(i)].width     = 9
        ws.column_dimensions[get_column_letter(i + 1)].width = 28


# ── Public entry point ────────────────────────────────────────────────────────

def build_spreadsheet(
    results:  List[TickerAnalysis],
    bulls:    List[Tuple],
    bears:    List[Tuple],
    universe: str,
    tag:      str = "",
) -> Path:
    """Build and save the full Excel workbook. Returns the file path."""
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M")
    stem = f"scan_{tag}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    path = config.OUTPUT_DIR / f"{stem}.xlsx"

    wb = openpyxl.Workbook()
    _sheet_top_picks(wb, bulls, bears, universe, ts)
    _sheet_all_results(wb, results, universe, ts)
    _sheet_signal_detail(wb, results)

    wb.save(path)
    log.info("Spreadsheet saved → %s", path)
    return path



# ── Sheet 4: Fibonacci Projections (Top 10 + full universe) ──────────────────

def _sheet_fibonacci(wb, results, bulls, bears) -> None:
    ws = wb.create_sheet("Fibonacci Projections")
    ws.sheet_view.showGridLines = False

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:U1")
    t = ws["A1"]
    t.value     = "Fibonacci Price Projections — Entry · Stop Loss · Targets T1/T2/T3 · Risk-Reward"
    t.font      = Font(name="Calibri", bold=True, size=12, color=HEADER_FG)
    t.fill      = _fill(HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Helper to write a picks sub-section ───────────────────────────────────
    def write_picks_section(picks, section_title, fill_color, start_row):
        ws.merge_cells(f"A{start_row}:U{start_row}")
        c = ws.cell(row=start_row, column=1, value=section_title)
        c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill      = _fill(fill_color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[start_row].height = 20
        start_row += 1

        cols = [
            "Rank", "Ticker", "Price", "Score", "Direction",
            "Anchor Type", "Swing High", "Swing Low", "Range",
            "Entry (38.2%)", "Stop (61.8%)",
            "Target 1 (100%)", "Target 2 (127.2%)", "Target 3 (161.8%)",
            "R/R T1", "R/R T2", "R/R T3",
            "Next-Hr Target", "Support", "Resistance", "Setup Quality"
        ]
        _write_header_row(ws, start_row, cols, ACCENT_BG)
        ws.row_dimensions[start_row].height = 30
        start_row += 1

        for rank, (ta, cs) in enumerate(picks, 1):
            fib = ta.fib
            dir_str = fib.direction.upper() if fib else "—"
            dfill = _fill(BULL_LIGHT) if (fib and fib.direction == "bullish") else                     _fill(BEAR_LIGHT) if (fib and fib.direction == "bearish") else                     _fill(NEUTRAL_BG)

            def fp(v):
                return round(v, 2) if v else None

            # Setup quality: based on conviction and R/R
            quality = "—"
            if fib and fib.risk_reward_t2 and cs.conviction_pct:
                rr = fib.risk_reward_t2
                pct = cs.conviction_pct
                if rr >= 3.0 and pct >= 70: quality = "A+ — High conviction"
                elif rr >= 2.0 and pct >= 55: quality = "A — Good setup"
                elif rr >= 1.5 and pct >= 40: quality = "B — Moderate"
                else: quality = "C — Low probability"

            row_vals = [
                rank,
                ta.ticker,
                fp(ta.price) if ta.price else None,
                ta.net_score,
                dir_str,
                fib.anchor_type                         if fib else "—",
                fp(fib.swing_high)                      if fib else None,
                fp(fib.swing_low)                       if fib else None,
                fp(fib.swing_range)                     if fib else None,
                fp(fib.entry_price)                     if fib else None,
                fp(fib.stop_loss)                       if fib else None,
                fp(fib.target_1)                        if fib else None,
                fp(fib.target_2)                        if fib else None,
                fp(fib.target_3)                        if fib else None,
                round(fib.risk_reward_t1, 2)            if (fib and fib.risk_reward_t1) else None,
                round(fib.risk_reward_t2, 2)            if (fib and fib.risk_reward_t2) else None,
                round(fib.risk_reward_t3, 2)            if (fib and fib.risk_reward_t3) else None,
                fp(fib.next_hour_target)                if fib else None,
                fp(fib.support_1)                       if fib else None,
                fp(fib.resistance_1)                    if fib else None,
                quality,
            ]

            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=start_row, column=c_idx, value=val)
                cell.font      = _body_font(bold=(c_idx in (1, 2)))
                cell.border    = _thin_border()
                cell.alignment = Alignment(
                    horizontal="center" if c_idx not in (2, 6, 21) else "left",
                    vertical="center"
                )
                # Color coding
                if c_idx == 4:   cell.fill = _score_fill(ta.net_score)
                elif c_idx == 10: cell.fill = _fill("E8F5E9")   # entry — light green
                elif c_idx == 11: cell.fill = _fill("FFEBEE")   # stop  — light red
                elif c_idx in (12, 13, 14):  cell.fill = _fill("FFF9C4")  # targets — yellow
                elif c_idx in (15, 16, 17):  # R/R — color by value
                    if isinstance(val, (int, float)):
                        if val >= 3.0:   cell.fill = _fill(BULL_LIGHT)
                        elif val >= 2.0: cell.fill = _fill("E8F5E9")
                        elif val >= 1.0: cell.fill = _fill(NEUTRAL_BG)
                        else:            cell.fill = _fill(BEAR_LIGHT)
                        cell.font = _body_font(bold=True)
                else:
                    cell.fill = dfill

            ws.row_dimensions[start_row].height = 20
            start_row += 1
        return start_row + 1

    # ── Write bullish and bearish top picks ───────────────────────────────────
    next_row = write_picks_section(bulls, "▲  TOP 5 BULLISH — Fibonacci Entry / Exit Plan", BULL_DARK, 3)
    next_row = write_picks_section(bears, "▼  TOP 5 BEARISH — Fibonacci Short / Exit Plan", BEAR_DARK, next_row)

    # ── All-universe Fibonacci summary ────────────────────────────────────────
    fib_results = [ta for ta in results if ta.fib]
    if fib_results:
        next_row += 1
        ws.merge_cells(f"A{next_row}:U{next_row}")
        c = ws.cell(row=next_row, column=1, value=f"📊 Full Universe Fibonacci Summary — {len(fib_results)} tickers")
        c.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
        c.fill      = _fill(HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[next_row].height = 20
        next_row += 1

        universe_cols = [
            "Ticker", "Price", "Score", "Direction", "Anchor",
            "Swing High", "Swing Low",
            "Entry (38.2%)", "Stop (61.8%)",
            "T1 (100%)", "T2 (127.2%)", "T3 (161.8%)",
            "R/R T1", "R/R T2", "R/R T3",
            "Next-Hr Target", "Support", "Resistance",
            "R 23.6%", "R 50.0%", "E 161.8%",
        ]
        _write_header_row(ws, next_row, universe_cols, ACCENT_BG)
        ws.row_dimensions[next_row].height = 28
        next_row += 1

        for i, ta in enumerate(fib_results):
            fib = ta.fib
            dfill = _fill(BULL_LIGHT) if fib.direction == "bullish" else                     _fill(BEAR_LIGHT) if fib.direction == "bearish" else                     _fill(NEUTRAL_BG)

            r_prices = {lvl.label: lvl.price for lvl in fib.retracements}
            e_prices = {lvl.label: lvl.price for lvl in fib.extensions}

            row_vals = [
                ta.ticker,
                round(ta.price, 2) if ta.price else None,
                ta.net_score,
                fib.direction.upper(),
                fib.anchor_type,
                fib.swing_high, fib.swing_low,
                round(fib.entry_price, 2)  if fib.entry_price else None,
                round(fib.stop_loss, 2)    if fib.stop_loss   else None,
                round(fib.target_1, 2)     if fib.target_1    else None,
                round(fib.target_2, 2)     if fib.target_2    else None,
                round(fib.target_3, 2)     if fib.target_3    else None,
                round(fib.risk_reward_t1, 2) if fib.risk_reward_t1 else None,
                round(fib.risk_reward_t2, 2) if fib.risk_reward_t2 else None,
                round(fib.risk_reward_t3, 2) if fib.risk_reward_t3 else None,
                round(fib.next_hour_target, 2) if fib.next_hour_target else None,
                round(fib.support_1, 2)    if fib.support_1    else None,
                round(fib.resistance_1, 2) if fib.resistance_1 else None,
                r_prices.get("R 23.6%"), r_prices.get("R 50.0%"), e_prices.get("E 161.8%"),
            ]

            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=next_row, column=c_idx, value=val)
                cell.font      = _body_font(bold=(c_idx == 1))
                cell.border    = _thin_border()
                cell.alignment = Alignment(
                    horizontal="center" if c_idx not in (1, 5) else "left",
                    vertical="center"
                )
                if c_idx == 3:    cell.fill = _score_fill(ta.net_score)
                elif c_idx == 8:  cell.fill = _fill("E8F5E9")
                elif c_idx == 9:  cell.fill = _fill("FFEBEE")
                elif c_idx in (10, 11, 12): cell.fill = _fill("FFF9C4")
                elif c_idx in (13, 14, 15):
                    if isinstance(val, (int, float)):
                        cell.fill = _fill(BULL_LIGHT) if val >= 2.0 else                                     _fill(NEUTRAL_BG) if val >= 1.0 else                                     _fill(BEAR_LIGHT)
                        cell.font = _body_font(bold=True)
                elif i % 2 == 0:
                    cell.fill = _fill("F9F9F9")
            next_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    width_map = {
        "A": 9, "B": 9, "C": 7, "D": 11, "E": 22,
        "F": 11, "G": 11, "H": 10, "I": 10,
        "J": 13, "K": 14, "L": 14,
        "M": 9, "N": 9, "O": 9,
        "P": 14, "Q": 11, "R": 11,
        "S": 10, "T": 10, "U": 10,
    }
    _set_col_widths(ws, width_map)
    ws.freeze_panes = "A3"


# ── build_spreadsheet (replaces the patched version) ─────────────────────────

def build_spreadsheet(results, bulls, bears, universe, tag=""):
    from datetime import datetime
    import openpyxl
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M")
    stem = f"scan_{tag}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    path = config.OUTPUT_DIR / f"{stem}.xlsx"

    wb = openpyxl.Workbook()
    _sheet_top_picks(wb, bulls, bears, universe, ts)
    _sheet_all_results(wb, results, universe, ts)
    _sheet_signal_detail(wb, results)
    _sheet_fibonacci(wb, results, bulls, bears)

    wb.save(path)
    log.info("Spreadsheet saved → %s", path)
    return path
